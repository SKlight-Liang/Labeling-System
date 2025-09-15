import json
import io
import csv
import sqlite3
from datetime import datetime
from flask import Response, jsonify


class DataExporter:
    """数据导出器类"""
    
    def __init__(self, db_path='users.db'):
        """
        初始化数据导出器
        
        Args:
            db_path: 数据库文件路径
        """
        self.db_path = db_path
        
        # 字段映射 - 包含所有字段
        self.field_mapping = {
            'id': 'ID',
            'question_cn': '中文问题',
            'answer_cn': '中文答案', 
            'question_en': '英文问题',
            'answer_en': '英文答案',
            'subject': '学科',
            'clue_urls': '线索URLs',
            'traces': '推理轨迹',
            'answer_clue': '答案线索',
            'answer_url': '答案URL',
            'check_info': '检查信息',
            'dfsw_info': 'DFSW信息',
            'created_at': '创建时间',
            'uploaded_by': '上传者',
            'first_annotator': '第一标注员',
            'first_annotation_result': '第一标注结果',
            'first_annotation_time': '第一标注时间',
            'second_annotator': '第二标注员',
            'second_annotation_result': '第二标注结果',
            'second_annotation_time': '第二标注时间',
            'final_status': '最终状态',
            'annotation_result': '管理员标注结果',
            'annotated_by': '管理员',
            'annotated_at': '标注时间'
        }

    def build_query(self, filters=None):
        """
        构建数据查询SQL
        
        Args:
            filters: 筛选条件字典
                - status: 'all', 'pending', 'first_completed', 'completed'
                - annotator: 'all' 或具体用户名
                - agreement: 'all', 'agreed', 'conflicted'
                - result: 'all', 'good', 'bad', 'uncertain', 'agreed_good', 'agreed_bad', 'agreed_uncertain', 'admin_approved', 'double_annotated_good', 'high_quality'
        
        Returns:
            tuple: (query_sql, query_params)
        """
        if filters is None:
            filters = {}
            
        query_conditions = []
        query_params = []
        
        # 状态筛选
        status = filters.get('status', 'all')
        if status != 'all':
            if status == 'pending':
                query_conditions.append("(first_annotator = '' OR first_annotator IS NULL) AND (second_annotator = '' OR second_annotator IS NULL)")
            elif status == 'first_completed':
                query_conditions.append("(first_annotator != '' AND first_annotator IS NOT NULL) AND (second_annotator = '' OR second_annotator IS NULL)")
            elif status == 'completed':
                query_conditions.append("(first_annotator != '' AND first_annotator IS NOT NULL) AND (second_annotator != '' AND second_annotator IS NOT NULL)")
        
        # 标注员筛选
        annotator = filters.get('annotator', 'all')
        if annotator != 'all':
            query_conditions.append("(first_annotator = ? OR second_annotator = ?)")
            query_params.extend([annotator, annotator])
        
        # 一致性筛选
        agreement = filters.get('agreement', 'all')
        if agreement != 'all':
            query_conditions.append("final_status = ?")
            query_params.append(agreement)
        
        # 结果筛选
        result = filters.get('result', 'all')
        if result != 'all':
            if result == 'good':
                query_conditions.append("(first_annotation_result = 'good' OR second_annotation_result = 'good' OR annotation_result = 'good')")
            elif result == 'bad':
                query_conditions.append("(first_annotation_result = 'bad' OR second_annotation_result = 'bad' OR annotation_result = 'bad')")
            elif result == 'uncertain':
                query_conditions.append("(first_annotation_result = 'uncertain' OR second_annotation_result = 'uncertain' OR annotation_result = 'uncertain')")
            elif result == 'agreed_good':
                query_conditions.append("first_annotation_result = 'good' AND second_annotation_result = 'good'")
            elif result == 'agreed_bad':
                query_conditions.append("first_annotation_result = 'bad' AND second_annotation_result = 'bad'")
            elif result == 'agreed_uncertain':
                query_conditions.append("first_annotation_result = 'uncertain' AND second_annotation_result = 'uncertain'")
            elif result == 'admin_approved':
                query_conditions.append("annotation_result = 'good'")
            elif result == 'high_quality':
                query_conditions.append("(annotation_result = 'good' OR (first_annotation_result = 'good' AND second_annotation_result = 'good'))")
        
        # 构建完整查询
        base_query = """SELECT id, question_cn, answer_cn, question_en, answer_en, subject, 
                               clue_urls, traces, answer_clue, answer_url, check_info, dfsw_info,
                               created_at, uploaded_by,
                               first_annotator, first_annotation_result, first_annotation_time,
                               second_annotator, second_annotation_result, second_annotation_time,
                               final_status, annotation_result, annotated_by, annotated_at FROM qa_data"""
        
        if query_conditions:
            base_query += " WHERE " + " AND ".join(query_conditions)
        
        base_query += " ORDER BY id"
        
        return base_query, query_params

    def get_data(self, filters=None):
        """
        从数据库获取数据
        
        Args:
            filters: 筛选条件字典
        
        Returns:
            list: 数据列表
        """
        query, params = self.build_query(filters)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(query, params)
        data = cursor.fetchall()
        conn.close()
        
        return data

    def export_to_csv(self, data, filename_prefix="data_export"):
        """
        导出为CSV格式
        
        Args:
            data: 数据列表
            filename_prefix: 文件名前缀
        
        Returns:
            Flask Response对象
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 写入表头
        headers = list(self.field_mapping.values())
        writer.writerow(headers)
        
        # 写入数据
        columns = list(self.field_mapping.keys())
        for row in data:
            csv_row = []
            for i, value in enumerate(row):
                if i < len(columns):
                    # 处理JSON字段，确保中文正确显示
                    if columns[i] in ['clue_urls', 'traces', 'check_info', 'dfsw_info']:
                        try:
                            if value:
                                # 解析JSON并重新序列化，确保中文正确显示
                                parsed_data = json.loads(value)
                                csv_row.append(json.dumps(parsed_data, ensure_ascii=False))
                            else:
                                csv_row.append('')
                        except (json.JSONDecodeError, TypeError):
                            csv_row.append(str(value) if value else '')
                    else:
                        csv_row.append(str(value) if value else '')
                else:
                    csv_row.append('')
            writer.writerow(csv_row)
        
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.csv"
        
        # 确保CSV文件支持UTF-8编码
        csv_content = output.getvalue()
        
        return Response(
            csv_content.encode('utf-8-sig'),  # 使用UTF-8 BOM确保Excel正确识别编码
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    def export_to_json(self, data, filename_prefix="data_export"):
        """
        导出为JSON格式
        
        Args:
            data: 数据列表
            filename_prefix: 文件名前缀
        
        Returns:
            Flask Response对象
        """
        columns = list(self.field_mapping.keys())
        
        json_data = []
        for row in data:
            item_dict = {}
            for i, value in enumerate(row):
                if i < len(columns):
                    # 处理JSON字段
                    if columns[i] in ['clue_urls', 'traces', 'check_info', 'dfsw_info']:
                        try:
                            # 解析JSON并重新序列化，确保中文正确显示
                            parsed_data = json.loads(value) if value else None
                            item_dict[columns[i]] = parsed_data
                        except (json.JSONDecodeError, TypeError):
                            item_dict[columns[i]] = value
                    else:
                        item_dict[columns[i]] = value
            json_data.append(item_dict)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{timestamp}.json"
        
        # 确保中文字符正确显示
        json_string = json.dumps(json_data, ensure_ascii=False, indent=2)
        
        return Response(
            json_string.encode('utf-8'),
            mimetype='application/json; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    def export_to_excel(self, data, filename_prefix="data_export"):
        """
        导出为Excel格式
        
        Args:
            data: 数据列表
            filename_prefix: 文件名前缀
        
        Returns:
            Flask Response对象或错误响应
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "数据导出"
            
            # 表头
            headers = list(self.field_mapping.values())
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
                # 设置列宽
                if col in [2, 3, 4, 5]:  # 问题和答案列
                    ws.column_dimensions[get_column_letter(col)].width = 30
                elif col in [7, 8, 9, 10, 11, 12]:  # URLs, traces, clue等信息列
                    ws.column_dimensions[get_column_letter(col)].width = 25
                else:
                    ws.column_dimensions[get_column_letter(col)].width = 15
            
            # 写入数据
            columns = list(self.field_mapping.keys())
            for row_num, row_data in enumerate(data, 2):
                for col, value in enumerate(row_data, 1):
                    # 处理JSON字段，确保中文正确显示
                    if col <= len(columns):
                        col_index = col - 1
                        if col_index < len(columns) and columns[col_index] in ['clue_urls', 'traces', 'check_info', 'dfsw_info']:
                            try:
                                if value:
                                    # 解析JSON并重新序列化，确保中文正确显示
                                    parsed_data = json.loads(value)
                                    display_value = json.dumps(parsed_data, ensure_ascii=False, indent=2)
                                else:
                                    display_value = ''
                            except (json.JSONDecodeError, TypeError):
                                display_value = str(value) if value else ''
                        else:
                            display_value = str(value) if value else ''
                    else:
                        display_value = str(value) if value else ''
                    
                    # 处理长文本换行
                    cell = ws.cell(row=row_num, column=col, value=display_value)
                    if col in [2, 3, 4, 5, 7, 8, 9, 10]:  # 文本内容列
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # 冻结首行
            ws.freeze_panes = "A2"
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename={filename}'}
            )
            
        except ImportError:
            return jsonify({'error': 'Excel导出功能需要安装openpyxl库。请使用CSV或JSON格式导出。'}), 400
        except Exception as e:
            return jsonify({'error': f'Excel导出失败: {str(e)}'}), 500

    def export_data(self, export_format='csv', filters=None, filename_prefix="data_export"):
        """
        统一的数据导出接口
        
        Args:
            export_format: 导出格式 ('csv', 'json', 'xlsx')
            filters: 筛选条件字典
            filename_prefix: 文件名前缀
        
        Returns:
            Flask Response对象
        """
        # 获取数据
        data = self.get_data(filters)
        
        # 根据格式导出
        if export_format == 'csv':
            return self.export_to_csv(data, filename_prefix)
        elif export_format == 'json':
            return self.export_to_json(data, filename_prefix)
        elif export_format == 'xlsx':
            return self.export_to_excel(data, filename_prefix)
        else:
            return jsonify({'error': '不支持的导出格式'}), 400

    def get_annotators(self):
        """
        获取标注员列表
        
        Returns:
            list: 标注员用户名列表
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT DISTINCT u.username 
            FROM users u 
            WHERE u.username IN (
                SELECT DISTINCT first_annotator FROM qa_data WHERE first_annotator != '' AND first_annotator IS NOT NULL
                UNION
                SELECT DISTINCT second_annotator FROM qa_data WHERE second_annotator != '' AND second_annotator IS NOT NULL
            )
            ORDER BY u.username
        """)
        
        annotators = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        return annotators

    def get_export_statistics(self, filters=None):
        """
        获取导出数据的统计信息
        
        Args:
            filters: 筛选条件字典
        
        Returns:
            dict: 统计信息
        """
        data = self.get_data(filters)
        
        total_count = len(data)
        if total_count == 0:
            return {'total_count': 0}
        
        # 统计各种状态
        stats = {
            'total_count': total_count,
            'completed_count': 0,
            'pending_count': 0,
            'first_completed_count': 0,
            'agreed_count': 0,
            'conflicted_count': 0,
            'good_count': 0,
            'bad_count': 0,
            'uncertain_count': 0
        }
        
        for row in data:
            # 根据字段位置统计（基于field_mapping的顺序）
            first_annotator = row[12]  # first_annotator
            second_annotator = row[15]  # second_annotator
            final_status = row[18]  # final_status
            first_result = row[13]  # first_annotation_result
            second_result = row[16]  # second_annotation_result
            admin_result = row[19]  # annotation_result
            
            # 状态统计
            if first_annotator and second_annotator:
                stats['completed_count'] += 1
            elif first_annotator:
                stats['first_completed_count'] += 1
            else:
                stats['pending_count'] += 1
            
            # 一致性统计
            if final_status == 'agreed':
                stats['agreed_count'] += 1
            elif final_status == 'conflicted':
                stats['conflicted_count'] += 1
            
            # 结果统计
            if first_result == 'good' or second_result == 'good' or admin_result == 'good':
                stats['good_count'] += 1
            if first_result == 'bad' or second_result == 'bad' or admin_result == 'bad':
                stats['bad_count'] += 1
            if first_result == 'uncertain' or second_result == 'uncertain' or admin_result == 'uncertain':
                stats['uncertain_count'] += 1
        
        return stats


# 创建全局导出器实例
data_exporter = DataExporter()

# 为了向后兼容，保留原有的函数接口
def export_to_csv(data, filename_prefix="data_export"):
    """向后兼容的CSV导出函数"""
    return data_exporter.export_to_csv(data, filename_prefix)

def export_to_json(data, filename_prefix="data_export"):
    """向后兼容的JSON导出函数"""
    return data_exporter.export_to_json(data, filename_prefix)

def export_to_excel(data, filename_prefix="data_export"):
    """向后兼容的Excel导出函数"""
    return data_exporter.export_to_excel(data, filename_prefix)